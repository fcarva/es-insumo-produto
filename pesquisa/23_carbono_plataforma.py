# -*- coding: utf-8 -*-
"""
23_carbono_plataforma.py — EXTENSÃO CARBONO-PLATAFORMA (insumo-produto ambientalmente
estendido, EEIO). Estima o CO2 INCORPORADO na produção/pauta do ES e a "fração que vaza"
(analogamente ao vazamento do multiplicador), reusando a MIP-ES-BR 2008.

>>> INTEGRIDADE: este script NÃO inventa emissões. Ele EXIGE um vetor real de intensidade
    de CO2 por setor (de SEEG / Haddad et al. 2025 / IBGE Contas Ambientais). Sem o dado,
    ele PARA e imprime a especificação do arquivo. Nada de número-fantasma. <<<

Método (Miller & Blair 2009, cap. ext. ambiental; Carvalho & Perobelli 2009):
  f_i = CO2_i / x_i                     intensidade direta (tCO2 por R$ mi)
  m   = f' B                            multiplicador de carbono (tCO2/R$ mi de dem. final)
  E   = f' B y                          CO2 incorporado na demanda final y
  vazamento de carbono_j = parcela de m_j emitida FORA do ES (linhas RB de f·B)
  carbono da pauta = sum_j (peso_export_j * m_j)   (analogia da upstreamness)
Hipótese declarada: tecnologia de emissão NACIONAL aplicada a ES e RB (como na camada CGV).

Entrada do dado: dados/co2_intensidade.csv com 26 linhas (ordem dos setores do ES) e
coluna 'co2_tco2_por_Rmi' (tCO2 por R$ milhão de produção, intensidade direta nacional do setor).
"""
import os, sys, csv
import numpy as np
import openpyxl
if hasattr(sys.stdout, "reconfigure"): sys.stdout.reconfigure(encoding="utf-8")

DATA = r"C:/Users/DELL/Downloads/Material IO/Matrizes/MIP-ES-BR (2008).xlsx"
CO2  = r"C:/Users/DELL/Documents/es-insumo-produto/dados/co2_intensidade.csv"
OUT  = r"C:/Users/DELL/Documents/es-insumo-produto/pesquisa/outputs"
num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0

def parse_mip():
    wb = openpyxl.load_workbook(DATA, data_only=True); ws = wb["ES-BR"]
    names = [str(ws.cell(r, 2).value).strip() for r in range(7, 33)]
    Z  = np.array([[num(ws.cell(r, c).value) for c in range(5, 57)] for r in range(7, 59)])
    FD = np.array([[num(ws.cell(r, c).value) for c in range(58, 70)] for r in range(7, 59)])
    vbp = np.array([num(ws.cell(80, c).value) for c in range(5, 57)])
    wb.close()
    return names, Z, FD, vbp

def main():
    if not os.path.exists(CO2):
        print("="*78)
        print("DADO DE CO2 AUSENTE — o script não inventa emissões. Forneça:")
        print(f"  arquivo: {CO2}")
        print("  formato: CSV com 26 linhas (ordem dos setores do ES, ver names abaixo) e colunas:")
        print("           setor,co2_tco2_por_Rmi")
        print("  fonte recomendada (uma das):")
        print("   1) Haddad et al. (2025) — satélite de CO2 (68 setores, 2019) -> agregar p/ 26")
        print("   2) SEEG (plataforma.seeg.eco.br) — emissões por atividade econômica x UF (ES)")
        print("   3) IBGE Contas Econômicas Ambientais — emissões atmosféricas por atividade")
        names, *_ = parse_mip()
        print("-"*78); print("ordem esperada dos 26 setores do ES (col 'setor'):")
        for i, nm in enumerate(names, 1): print(f"   {i:2d}. {nm}")
        print("="*78)
        sys.exit(2)

    names, Z, FD, vbp = parse_mip()
    n = 52; L = np.arange(0, 26); M = np.arange(26, 52)
    x = np.where(vbp == 0, Z.sum(1), vbp); xs = np.where(x == 0, 1.0, x)
    A = Z / xs[None, :]; B = np.linalg.inv(np.eye(n) - A)

    # vetor de CO2 (intensidade direta nacional, tCO2/R$mi) — aplicado a ES e RB
    f26 = {}
    with open(CO2, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            f26[row["setor"].strip()] = float(row["co2_tco2_por_Rmi"])
    f = np.array([f26.get(nm, np.nan) for nm in names])
    if np.isnan(f).any():
        faltam = [names[i] for i in np.where(np.isnan(f))[0]]
        print("CSV incompleto — faltam setores:", faltam); sys.exit(2)
    f52 = np.concatenate([f, f])                       # tecnologia nacional p/ ES e RB

    m = f52 @ B                                        # multiplicador de carbono por setor-coluna
    # vazamento de carbono: parcela de m_j emitida fora do ES (linhas RB)
    m_in = (f52[L, None] * B[L, :]).sum(0)
    leak_c = 1 - m_in / np.where(m == 0, 1, m)
    # pauta exportadora do ES (Export = 1a categoria da demanda final do ES) -> carbono da pauta
    exp_L = FD[L, 0]                                   # exportações do ES por setor
    w_exp = exp_L / exp_L.sum() if exp_L.sum() else np.zeros_like(exp_L)
    carbono_pauta = float((w_exp * m[L]).sum())
    # CO2 emitido na produção do ES (production-based) e incorporado na dem. final do ES
    co2_prod_ES = float((f[L] * x[L]).sum())
    yL = FD[L, 0:6].sum(1)                             # dem. final do ES por produtos do ES
    co2_emb_demES = float(f52 @ B @ np.concatenate([yL, np.zeros(26)]))

    print("="*72)
    print("EXTENSÃO CARBONO-PLATAFORMA — ES (2008), tecnologia de emissão nacional")
    print("="*72)
    print(f"  multiplicador de carbono médio (26 setores ES): {m[L].mean():.1f} tCO2/R$mi")
    print(f"  vazamento de carbono médio (emitido fora do ES): {leak_c[L].mean():.1%}")
    print(f"  carbono da PAUTA exportadora do ES: {carbono_pauta:.1f} tCO2/R$mi")
    print(f"  CO2 emitido na produção do ES (production-based): {co2_prod_ES:,.0f} tCO2")
    print(f"  CO2 incorporado na demanda final do ES: {co2_emb_demES:,.0f} tCO2")
    o = np.argsort(m[L])[::-1]
    print("  top-5 setores por multiplicador de carbono:")
    for j in o[:5]: print(f"    {names[L[j]][:30]:30s} {m[L][j]:.1f} tCO2/R$mi  (vaza {leak_c[L][j]:.0%})")
    print("="*72)
    with open(os.path.join(OUT, "carbono_plataforma.csv"), "w", newline="", encoding="utf-8") as fo:
        w = csv.writer(fo); w.writerow(["setor","mult_carbono_tco2_Rmi","vazamento_carbono","intensidade_direta"])
        for i, j in enumerate(L): w.writerow([names[i], f"{m[j]:.3f}", f"{leak_c[j]:.4f}", f"{f[j]:.3f}"])
    print("salvo: carbono_plataforma.csv")

if __name__ == "__main__":
    main()
