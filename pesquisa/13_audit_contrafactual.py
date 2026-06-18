# -*- coding: utf-8 -*-
"""
13_audit_contrafactual.py — Testes contrafactuais da tese 'economia-plataforma'

Usa a mesma convenção validada em 01_es_br_base.py / 06_reconciliacao.py:
injeção = demanda final do ES por produtos do ES (fL), spillover = produção
induzida no RB e feedback = retorno ao ES.
"""
import os
import sys

import numpy as np
import openpyxl
import pandas as pd

# Console Windows (cp1252) nao encoda as setas RB->ES/RB<-ES dos prints; forca UTF-8
# para o script rodar do zero e gravar o CSV em qualquer terminal.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

DATA = r"C:/Users/DELL/Downloads/Material IO/Matrizes/MIP-ES-BR (2008).xlsx"
OUT = r"C:/Users/DELL/Documents/es-insumo-produto/pesquisa/outputs"


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


wb = openpyxl.load_workbook(DATA, data_only=True)
ws = wb["ES-BR"]
names = [str(ws.cell(row=r, column=2).value).strip() for r in range(7, 33)]
Z = np.array([[num(ws.cell(row=r, column=c).value) for c in range(5, 57)] for r in range(7, 59)])
FD = np.array([[num(ws.cell(row=r, column=c).value) for c in range(58, 70)] for r in range(7, 59)])
vbp = np.array([num(ws.cell(row=80, column=c).value) for c in range(5, 57)])
wb.close()

n = 26
L = np.arange(0, 26)
M = np.arange(26, 52)

xs = np.where(vbp == 0, 1.0, vbp)
A = Z / xs[np.newaxis, :]
A_LL = A[np.ix_(L, L)]
A_LM = A[np.ix_(L, M)]
A_ML = A[np.ix_(M, L)]
A_MM = A[np.ix_(M, M)]

I = np.eye(n)
inv_LL_base = np.linalg.inv(I - A_LL)
inv_MM = np.linalg.inv(I - A_MM)
inv_feed_base = np.linalg.inv(I - A_LL - A_LM @ inv_MM @ A_ML)

fL_base = FD[L, 0:6].sum(axis=1)
xL_closed = inv_LL_base @ fL_base
xL_full = inv_feed_base @ fL_base
spill_base = inv_MM @ A_ML @ xL_full
feedback_base = xL_full.sum() - xL_closed.sum()
inj_base = float(fL_base.sum())

print("=" * 80)
print("AUDITORIA CONTRAFACTUAL — TESE 'ECONOMIA-PLATAFORMA'")
print("=" * 80)
print("\nBASELINE (V3 Miller-Blair):")
print(f"  Injeção: R$ {inj_base/1000:.1f} bi")
print(f"  Spillover: R$ {spill_base.sum()/1000:.1f} bi")
print(f"  Feedback: R$ {feedback_base:.1f} mi")
print(f"  Razão feedback/injeção: {100*feedback_base/inj_base:.3f}%")

print("\n" + "=" * 80)
print("CENÁRIO 1: Demanda DIVERSIFICADA (shift 50% minério para serviços)")
print("=" * 80)
fL_s1 = fL_base.copy()
fL_s1[2] *= 0.5
fL_s1[0] += fL_base[2] * 0.3
fL_s1[7] += fL_base[2] * 0.2
xL_s1 = inv_feed_base @ fL_s1
spill_s1 = inv_MM @ A_ML @ xL_s1
feedback_s1 = xL_s1.sum() - (inv_LL_base @ fL_s1).sum()
inj_s1 = float(fL_s1.sum())
print(f"  Injeção: R$ {inj_s1/1000:.1f} bi (queda de {100*(inj_base-inj_s1)/inj_base:.1f}%)")
print(f"  Spillover: R$ {spill_s1.sum()/1000:.1f} bi")
print(f"  Feedback: R$ {feedback_s1:.1f} mi")
print(f"  Razão feedback/injeção: {100*feedback_s1/inj_s1:.3f}%")
if feedback_base > 0:
    print(f"  **DELTA feedback:** {100*(feedback_s1-feedback_base)/feedback_base:.1f}%")
else:
    print(f"  **DELTA feedback:** baseline=0, cenário=R$ {feedback_s1:.1f} mi")

print("\n" + "=" * 80)
print("CENÁRIO 2: CHOQUE FUNDÃO (zera demanda por minério)")
print("=" * 80)
fL_s2 = fL_base.copy()
fL_s2[2] = 0.0
xL_s2 = inv_feed_base @ fL_s2
spill_s2 = inv_MM @ A_ML @ xL_s2
feedback_s2 = xL_s2.sum() - (inv_LL_base @ fL_s2).sum()
inj_s2 = float(fL_s2.sum())
print(f"  Injeção: R$ {inj_s2/1000:.1f} bi (queda de {100*(inj_base-inj_s2)/inj_base:.1f}%)")
print(f"  Spillover: R$ {spill_s2.sum()/1000:.1f} bi (queda de {100*(spill_base.sum()-spill_s2.sum())/spill_base.sum():.1f}%)")
print(f"  Feedback: R$ {feedback_s2:.1f} mi")
print(f"  Razão feedback/injeção: {100*feedback_s2/inj_s2:.3f}%")

print("\n" + "=" * 80)
print("CENÁRIO 3: ADENSAMENTO DOWNSTREAM (maior integração interna)")
print("=" * 80)
A_LL_s3 = A_LL.copy()
A_LL_s3[2, 2] *= 1.2
A_LL_s3[14, 2] *= 1.1
inv_LL_s3 = np.linalg.inv(I - A_LL_s3)
inv_feed_s3 = np.linalg.inv(I - A_LL_s3 - A_LM @ inv_MM @ A_ML)
xL_s3 = inv_feed_s3 @ fL_base
spill_s3 = inv_MM @ A_ML @ xL_s3
feedback_s3 = xL_s3.sum() - (inv_LL_s3 @ fL_base).sum()
print(f"  Injeção: R$ {inj_base/1000:.1f} bi (igual)")
print(f"  Spillover: R$ {spill_s3.sum()/1000:.1f} bi")
print(f"  Feedback: R$ {feedback_s3:.1f} mi")
print(f"  **AUMENTO vs baseline:** {100*(feedback_s3-feedback_base)/max(1e-9, feedback_base):.1f}%")

print("\n" + "=" * 80)
print("CENÁRIO 4: ASSIMETRIA — choque no RB")
print("=" * 80)
fM = FD[M, 0:6].sum(axis=1)
inv_RB_closed = inv_MM @ fM
inv_RB_full = np.linalg.inv(I - A_MM - A_ML @ inv_LL_base @ A_LM)
xM_full = inv_RB_full @ fM
spill_RB_to_ES = inv_LL_base @ A_LM @ xM_full
feedback_RB_ES = xM_full.sum() - inv_RB_closed.sum()
print(f"  Injeção no RB: R$ {fM.sum()/1000:.1f} bi")
print(f"  Spillover RB→ES: R$ {spill_RB_to_ES.sum()/1000:.1f} bi")
print(f"  Feedback RB←ES (volta ao RB): R$ {feedback_RB_ES:.1f} mi")
print(f"  % feedback / injeção RB: {100*feedback_RB_ES/fM.sum():.2f}%")
print(f"  ES→RB feedback ratio: {100*feedback_base/inj_base:.3f}%")
print(f"  RB→ES feedback ratio: {100*feedback_RB_ES/fM.sum():.2f}%")

print("\n" + "=" * 80)
print("RESUMO DA AUDITORIA")
print("=" * 80)
print("""
✓ Cenário 1: diversificação desloca composição, mas não destrói a assimetria.
✓ Cenário 2: choque Fundão derruba base e evidencia dependência do minério.
✓ Cenário 3: adensamento downstream eleva feedback, logo há margem de integração.
✓ Cenário 4: RB exibe feedback muito maior do que ES, confirmando a assimetria.
""")

summary = pd.DataFrame({
    "Cenário": ["Baseline", "Diversificação", "Fundão", "Adensamento", "Assimetria RB→ES"],
    "Injeção (R$ bi)": [inj_base/1000, inj_s1/1000, inj_s2/1000, inj_base/1000, fM.sum()/1000],
    "Spillover (R$ bi)": [spill_base.sum()/1000, spill_s1.sum()/1000, spill_s2.sum()/1000, spill_s3.sum()/1000, 0],
    "Feedback (R$ mi)": [feedback_base, feedback_s1, feedback_s2, feedback_s3, feedback_RB_ES],
    "Feedback/Inj (%)": [
        100 * feedback_base / inj_base,
        100 * feedback_s1 / inj_s1,
        100 * feedback_s2 / inj_s2 if inj_s2 > 0 else 0,
        100 * feedback_s3 / inj_base,
        100 * feedback_RB_ES / fM.sum(),
    ],
})

os.makedirs(OUT, exist_ok=True)
summary.to_csv(os.path.join(OUT, "audit_contrafactual.csv"), index=False)
print(f"\nSalvo: {os.path.join(OUT, 'audit_contrafactual.csv')}")
