# -*- coding: utf-8 -*-
"""
14_regionaliza_2015.py — constrói a MIP bi-regional ES x Resto do Brasil (2015)
por regionalização CILQ (Guilhoto & Sesso Filho, 2005), ancorada nos dados REAIS
de ES (TRU/MIP própria) e no residual nacional-menos-ES, com concordância
setorial 68->35.

Fontes (planilhas .xlsm de Celso Bissoli Sessa, CECEG/UFES, recebidas jun/2026;
motor de macro completo TRU->MIP->multiplicadores):
  - NAC: MIPBR_2015_Nivel_68.xlsm — TRU/MIP nacional 2015, nivel 68
  - ES : ESPIRITO_SANTO_2015.xlsm — TRU/MIP Espirito Santo 2015, nivel 35

Em ambos, aba '12' ("Consumo Intermediario de Origem Domestica") e a matriz Z
quadrada (atividade x atividade, R$ milhoes), com coluna 'Total do Produto'
(=x) e bloco 'Demanda Final' (Export[+Regional p/ ES], Governo, ISFLSF,
Familias, FBCF, Var.Estoque) logo apos; aba '01' linha 285 (NAC) / 185 (ES) =
"Fator Trabalho (Ocupacoes)", pessoal ocupado por atividade.

Por que regionalizar (e nao so subtrair): nenhuma das duas planilhas separa,
dentro do consumo intermediario "de origem domestica", o que e suprido de
dentro da propria regiao do que e suprido da outra regiao do Brasil. O
residual nacional-ES da a estrutura tecnologica do resto do Brasil, mas nao
os blocos cruzados Z_LM / Z_ML exigidos pelo modelo de Isard
(spillover/feedback). Aplica-se entao CILQ (cross-industry location quotient)
sobre os dois "totais domesticos" (Z_ES real e Z_RB residual) para estimar a
fracao local vs. importada de cada coeficiente.

Metodo CILQ (Guilhoto & Sesso Filho, 2005):
  LQ_i^r   = (x_i^r / X^r) / (x_i^BR / X^BR)                 [quociente locacional simples]
  CILQ_ij^r = LQ_i^r / LQ_j^r                                [cross-industry, p/ uso intermediario]
  RPC_ij^r  = min(CILQ_ij^r, 1)                              [coef. de compra regional]
  Z_local^r[i,j]    = RPC_ij^r * Z_total^r[i,j]
  Z_importado^r[i,j] = (1 - RPC_ij^r) * Z_total^r[i,j]
Para demanda final (sem "setor comprador", usa-se o SLQ sem termo cruzado):
  RPC_i^r = min(LQ_i^r, 1)

Resultado: Z empilhado 70x70 = [[Z_LL, Z_LM], [Z_ML, Z_MM]], x (70), ocup (70),
seguindo o layout esperado por src/io_core.py (codigo de setor + sufixo _L/_M).
"""
import os
import csv
import numpy as np
import openpyxl

UPLOAD_DIR = "/root/.claude/uploads/edbcb6fb-9812-5945-8b9d-66fc55f3c08c"
NAC_PATH = os.path.join(UPLOAD_DIR, "b26b2bac-MIPBR_2015_N_vel_68.xlsm")
ES_PATH = os.path.join(UPLOAD_DIR, "9810970a-ESP_RITO_SANTO_2015.xlsm")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DADOS = os.path.join(ROOT, "dados")
OUT = os.path.join(ROOT, "pesquisa", "outputs")
os.makedirs(OUT, exist_ok=True)

N_NAC = 68
N_ES = 35
EMP_ROW_NAC = 285
EMP_ROW_ES = 185
# colunas (1-indexadas) do bloco "Demanda Final" domestica (sem export "regional"):
# ordem: Exportacao(internacional), Governo, ISFLSF, Familias, FBCF, VarEstoque
FD_COLS_NAC = [74, 75, 76, 77, 78, 79]
FD_COLS_ES = [41, 43, 44, 45, 46, 47]
FD_REGIONAL_COL_ES = 42  # "Exportacao Regional" — usado so para auditoria
# x = "Demanda Total" (= "Total do Produto" [intermediario] + soma da "Demanda Final"
# completa, incl. Exportacao Regional na planilha ES) — e o vetor de PRODUCAO TOTAL
# correto; "Total do Produto" sozinho e so o consumo intermediario (~0 p/ setores
# nao-mercantis como administracao publica, saude/educacao publica).
DT_COL_NAC = 80
DT_COL_ES = 48

# --------------------------------------------------------------------------- #
# Concordancia setorial 68 -> 35 (mesmo namespace de codigo IBGE 4 digitos;
# mapeamento construido por correspondencia de nome de atividade — ver notas
# nas linhas marcadas "[julgamento]" para os casos sem correspondencia 1:1
# exata no rotulo do agregado de 35 setores).
# --------------------------------------------------------------------------- #
CONCORDANCIA = {
    "0191": ["0191"],
    "0192": ["0192"],
    "0280": ["0280"],
    "0580": ["0580", "0792"],                       # ES funde carvao+nao-metalicos com metalicos nao-ferrosos
    "0680": ["0680"],
    "0791": ["0791"],
    "1000": ["1091", "1092", "1093", "1100", "1200"],  # [julgamento] fumo (1200) sem rotulo proprio em ES; alocado em "Alimentos e bebidas"
    "1300": ["1300", "1400", "1500"],
    "1600": ["1600", "3180"],
    "1700": ["1700"],
    "1900": ["1991", "1992"],
    "2000": ["2091", "2092", "2093", "2100", "2200"],  # [julgamento] farmoquimicos/farmaceutica (2100) alocado em "quimicos"
    "2300": ["2300"],
    "2400": ["2491", "2492"],
    "2500": ["2500", "2600", "2700", "2800", "3300"],  # [julgamento] manutencao/reparacao de maquinas (3300) alocado aqui
    "2900": ["2991", "2992", "3000"],
    "3500": ["3500", "3680"],
    "4180": ["4180"],
    "4500": ["4500", "4680"],
    "4900": ["4900", "5000", "5100"],
    "5280": ["5280"],
    "5500": ["5500", "5600"],
    "5800": ["5800", "5980", "6100", "6280", "1800"],  # [julgamento] impressao/reproducao de gravacoes (1800) alocado em "informacao"
    "6480": ["6480"],
    "6800": ["6800"],
    "6900": ["6980", "7180", "7380"],
    "7800": ["7700", "7880", "8000"],
    "8400": ["8400"],
    "8591": ["8591"],
    "8592": ["8592"],
    "8691": ["8691"],
    "8692": ["8692"],
    "9080": ["9080"],
    "9480": ["9480"],
    "9700": ["9700"],
}


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def extrai(path, n, emp_row, fd_cols, dt_col, fd_regional_col=None):
    """Le aba '12' (Z, x, FD) e aba '01' (ocupacoes) de uma planilha Celso/CECEG."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws12 = wb["12"]
    codes, names, z_rows, x, fd_rows, fd_regional = [], [], [], [], [], []
    for r in range(7, 7 + n):
        row = list(ws12.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        codes.append(str(row[1]).strip())
        names.append(str(row[2]).replace("\n", " ").strip())
        z_rows.append([num(row[c - 1]) for c in range(5, 5 + n)])
        x.append(num(row[dt_col - 1]))  # "Demanda Total" = producao total (x)
        fd_rows.append([num(row[c - 1]) for c in fd_cols])
        if fd_regional_col:
            fd_regional.append(num(row[fd_regional_col - 1]))

    ws01 = wb["01"]
    emp_full = list(ws01.iter_rows(min_row=emp_row, max_row=emp_row, values_only=True))[0]
    ocup = [num(v) for v in emp_full[4:4 + n]]
    wb.close()
    return {
        "codes": codes,
        "names": names,
        "Z": np.array(z_rows),
        "x": np.array(x),
        "FD": np.array(fd_rows),                              # [Export_intl, Gov, ISFLSF, Familias, FBCF, VarEst]
        "FD_regional": np.array(fd_regional) if fd_regional else None,
        "ocup": np.array(ocup),
    }


def agrega_68_35(nac, concord, codes_es):
    """Agrega NAC (68) para a malha de 35 setores de ES, usando a concordancia."""
    idx = {c: i for i, c in enumerate(nac["codes"])}
    n = len(codes_es)
    Z35 = np.zeros((n, n))
    x35 = np.zeros(n)
    FD35 = np.zeros((n, nac["FD"].shape[1]))
    ocup35 = np.zeros(n)
    grupo_idx = [[idx[c] for c in concord[es_code]] for es_code in codes_es]
    for a, gi in enumerate(grupo_idx):
        x35[a] = nac["x"][gi].sum()
        FD35[a] = nac["FD"][gi].sum(axis=0)
        ocup35[a] = nac["ocup"][gi].sum()
        for b, gj in enumerate(grupo_idx):
            Z35[a, b] = nac["Z"][np.ix_(gi, gj)].sum()
    return Z35, x35, FD35, ocup35


def _lq(x_r, X_r, x_br, X_br):
    """LQ_i = (x_i^r/X^r)/(x_i^BR/X^BR); setores com x_i^BR=0 (ex.: '9700' Servicos
    domesticos, sem producao de mercado em nenhuma regiao) dao 0/0 -> tratados como 0
    (RPC=0 nesses casos e inofensivo: o Z/FD correspondente tambem e 0)."""
    lq = np.divide(x_r / X_r, x_br / X_br, out=np.zeros_like(x_r, dtype=float), where=(x_br != 0))
    return np.nan_to_num(lq, nan=0.0, posinf=0.0, neginf=0.0)


def cilq_rpc(x_r, X_r, x_br, X_br):
    """RPC_ij = min(LQ_i/LQ_j, 1) (CILQ) — matriz n x n para uso intermediario."""
    lq = _lq(x_r, X_r, x_br, X_br)
    cilq = np.divide(lq[:, None], lq[None, :], out=np.zeros((len(lq), len(lq))), where=(lq[None, :] != 0))
    return np.minimum(cilq, 1.0)


def slq_rpc(x_r, X_r, x_br, X_br):
    """RPC_i = min(LQ_i, 1) (SLQ, sem termo cruzado) — vetor n, p/ demanda final."""
    return np.minimum(_lq(x_r, X_r, x_br, X_br), 1.0)


def main():
    nac = extrai(NAC_PATH, N_NAC, EMP_ROW_NAC, FD_COLS_NAC, DT_COL_NAC)
    es = extrai(ES_PATH, N_ES, EMP_ROW_ES, FD_COLS_ES, DT_COL_ES, FD_REGIONAL_COL_ES)

    assert sorted(sum(CONCORDANCIA.values(), [])) == sorted(nac["codes"]), \
        "concordancia nao cobre exatamente os 68 codigos nacionais"
    assert list(CONCORDANCIA.keys()) == es["codes"], \
        "concordancia nao alinhada com os 35 codigos de ES"

    Z_nac35, x_nac35, FD_nac35, ocup_nac35 = agrega_68_35(nac, CONCORDANCIA, es["codes"])

    # ----- resto do Brasil = nacional(35) - ES, por residuo ----- #
    Z_rb = Z_nac35 - es["Z"]
    x_rb = x_nac35 - es["x"]
    ocup_rb = ocup_nac35 - es["ocup"]
    FD_rb = FD_nac35.copy()
    FD_rb[:, 0] = FD_nac35[:, 0] - es["FD"][:, 0]   # export internacional: BR - ES = RB
    FD_rb[:, 1:] = FD_nac35[:, 1:] - es["FD"][:, 1:]  # Gov/ISFLSF/Familias/FBCF/VarEst

    # ruido de residuo: Z_rb tem coeficientes >=0 por construcao economica (consumo
    # intermediario nao e negativo); pequenas violacoes vem de arredondamento/des-
    # sincronia entre as duas planilhas (fontes/momentos de fechamento distintos) e
    # sao zeradas, com a magnitude reportada na auditoria. FD_rb[:,4] (FBCF) e
    # FD_rb[:,5] (Variacao de Estoque) PODEM legitimamente ser negativos (desinves-
    # timento/desestocagem) e não são tocados.
    n_clip_z = int((Z_rb < 0).sum())
    clip_mag_z = float(-Z_rb[Z_rb < 0].sum()) if n_clip_z else 0.0
    Z_rb = np.maximum(Z_rb, 0.0)
    n_clip_fd = int((FD_rb[:, :4] < 0).sum())
    clip_mag_fd = float(-FD_rb[:, :4][FD_rb[:, :4] < 0].sum()) if n_clip_fd else 0.0
    FD_rb[:, :4] = np.maximum(FD_rb[:, :4], 0.0)

    checks = {
        "x_rb_min": float(x_rb.min()),
        "ocup_rb_min": float(ocup_rb.min()),
    }

    X_es, X_rb = es["x"].sum(), x_rb.sum()
    X_br = X_es + X_rb

    # ----- CILQ sobre consumo intermediario (Z) ----- #
    RPC_es = cilq_rpc(es["x"], X_es, x_nac35, X_br)
    RPC_rb = cilq_rpc(x_rb, X_rb, x_nac35, X_br)

    Z_LL = RPC_es * es["Z"]
    Z_ML = (1.0 - RPC_es) * es["Z"]       # ES compra de RB
    Z_MM = RPC_rb * Z_rb
    Z_LM = (1.0 - RPC_rb) * Z_rb          # RB compra de ES

    # ----- SLQ sobre demanda final domestica (Gov/ISFLSF/Familias/FBCF/VarEst) ----- #
    slq_es = slq_rpc(es["x"], X_es, x_nac35, X_br)
    slq_rb = slq_rpc(x_rb, X_rb, x_nac35, X_br)
    fd_local_es = slq_es[:, None] * es["FD"][:, 1:]
    fd_local_rb = slq_rb[:, None] * FD_rb[:, 1:]

    # y_L: demanda final por producao PROPRIA do ES (injecao do experimento
    # spillover/feedback) = demanda domestica suprida localmente + export
    # internacional do ES (exportacao e, por definicao, producao de origem ES).
    y_L = fd_local_es.sum(axis=1) + es["FD"][:, 0]
    y_M = fd_local_rb.sum(axis=1) + FD_rb[:, 0]

    # ----- auditoria: Z_LM (estimado) vs "Exportacao Regional" observada (ES) ----- #
    exportacao_regional_estimada = Z_LM.sum(axis=1)  # ES vende a RB (uso intermediario)
    diff_audit = exportacao_regional_estimada - es["FD_regional"]

    # ----- monta Z empilhado 70x70, x, ocup ----- #
    codes_L = [c + "_L" for c in es["codes"]]
    codes_M = [c + "_M" for c in es["codes"]]
    codigos = codes_L + codes_M
    n = N_ES
    Z_full = np.zeros((2 * n, 2 * n))
    Z_full[:n, :n] = Z_LL
    Z_full[:n, n:] = Z_LM
    Z_full[n:, :n] = Z_ML
    Z_full[n:, n:] = Z_MM
    x_full = np.concatenate([es["x"], x_rb])
    ocup_full = np.concatenate([es["ocup"], ocup_rb])

    # ----- grava dados/mip_es_br_2015.csv (Z + x + ocup, mesmo schema do 2008) ----- #
    mip_path = os.path.join(DADOS, "mip_es_br_2015.csv")
    with open(mip_path, "w", newline="", encoding="utf-8") as f:
        wcsv = csv.writer(f)
        wcsv.writerow(["setor"] + codigos + ["x", "ocup"])
        for i, code in enumerate(codigos):
            wcsv.writerow([code] + [f"{v:.6f}" for v in Z_full[i]] +
                          [f"{x_full[i]:.6f}", f"{ocup_full[i]:.6f}"])
    print("salvo:", mip_path)

    # ----- grava dados/concordancia_setorial_68_35.csv ----- #
    conc_path = os.path.join(DADOS, "concordancia_setorial_68_35.csv")
    with open(conc_path, "w", newline="", encoding="utf-8") as f:
        wcsv = csv.writer(f)
        wcsv.writerow(["codigo_es_35", "nome_es_35", "codigo_nacional_68"])
        nomes_es = dict(zip(es["codes"], es["names"]))
        for es_code, nac_codes in CONCORDANCIA.items():
            for nc in nac_codes:
                wcsv.writerow([es_code, nomes_es[es_code], nc])
    print("salvo:", conc_path)

    # ----- grava dados/fd_es_br_2015.csv (demanda final por bloco, p/ spillover/feedback) ----- #
    fd_path = os.path.join(DADOS, "fd_es_br_2015.csv")
    with open(fd_path, "w", newline="", encoding="utf-8") as f:
        wcsv = csv.writer(f)
        wcsv.writerow(["setor", "fd_local", "fd_importada_outra_regiao", "export_internacional",
                       "y_total_injecao", "exportacao_regional_observada_ES",
                       "exportacao_regional_estimada_Z_LM"])
        for i, code in enumerate(es["codes"]):
            wcsv.writerow([code + "_L", f"{fd_local_es[i].sum():.6f}",
                          f"{(es['FD'][i, 1:].sum() - fd_local_es[i].sum()):.6f}",
                          f"{es['FD'][i, 0]:.6f}", f"{y_L[i]:.6f}",
                          f"{es['FD_regional'][i]:.6f}", f"{exportacao_regional_estimada[i]:.6f}"])
        for i, code in enumerate(es["codes"]):
            wcsv.writerow([code + "_M", f"{fd_local_rb[i].sum():.6f}",
                          f"{(FD_rb[i, 1:].sum() - fd_local_rb[i].sum()):.6f}",
                          f"{FD_rb[i, 0]:.6f}", f"{y_M[i]:.6f}", "", ""])
    print("salvo:", fd_path)

    # ----- relatorio de auditoria ----- #
    print("=" * 70)
    print("AUDITORIA — MIP bi-regional ES x Resto do Brasil 2015 (CILQ)")
    print("=" * 70)
    print(f"Dimensao: {Z_full.shape}  (35 setores x 2 regioes)")
    print(f"Resto do Brasil minimos (devem ser >=0, ES e subconjunto do Brasil):")
    for k, v in checks.items():
        flag = "OK" if v >= -1e-6 else "** NEGATIVO **"
        print(f"    {k:14s} = {v:14,.2f}  [{flag}]")
    print(f"Ruido de residuo zerado — Z_rb: {n_clip_z} celula(s), R$ {clip_mag_z:,.2f} mi "
          f"(de {Z_rb.size}); FD_rb[Export/Gov/ISFLSF/Familias]: {n_clip_fd} celula(s), "
          f"R$ {clip_mag_fd:,.2f} mi")
    A_full = Z_full / np.where(x_full == 0, 1.0, x_full)[None, :]
    print(f"Soma de A por coluna — min/max: {A_full.sum(0).min():.3f} / {A_full.sum(0).max():.3f}  (deve ser <1)")
    print(f"Z_full >= 0 ? {bool((Z_full >= -1e-6).all())}")
    print("-" * 70)
    print("Auditoria CILQ: Exportacao Regional observada (ES, planilha) vs.")
    print("                Z_LM estimado (soma de vendas de ES p/ uso intermediario em RB)")
    print(f"    correlacao        : {np.corrcoef(exportacao_regional_estimada, es['FD_regional'])[0,1]:.3f}")
    print(f"    diff media        : {diff_audit.mean():14,.2f}")
    print(f"    diff abs. media   : {np.abs(diff_audit).mean():14,.2f}")
    print(f"    [nota] Z_LM so capta uso intermediario; 'Exportacao Regional' tambem inclui")
    print(f"           demanda final de RB por bens do ES (nao decomposta nesta planilha) —")
    print(f"           divergencia esperada, nao e erro de calculo.")
    print("=" * 70)


if __name__ == "__main__":
    main()
