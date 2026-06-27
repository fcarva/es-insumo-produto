# -*- coding: utf-8 -*-
"""
17_caracterizacao_es_2008.py — C1: RETRATO ESTRUTURAL do ES (2008) pela otica
insumo-produto, no formato do genero brasileiro (Figueiredo-Barros-Guilhoto MT;
Porsse-Peixoto-Palermo RS). Bateria completa para os 26 setores do ES:

  - multiplicadores de PRODUCAO, EMPREGO e RENDA  (tipo I e tipo II)
  - geracao direta/indireta/induzida de emprego e renda por R$ 1 mi de dem. final
  - ligacoes de Rasmussen-Hirschman (para tras: Leontief; para frente: Ghosh)
  - ligacoes PURAS (Guilhoto-Sonis-Hewings): PBL/PFL/PTL padronizadas
  - coeficiente de variacao das ligacoes (grau de integracao)
  - setores-chave (RH e ligacoes puras)

Tipo II = modelo fechado p/ as familias DO ES (renda das remuneracoes dos setores
do ES; consumo das familias do ES). Fonte: MIP-ES-BR (2008).xlsx, aba 'ES-BR'.
Layout: Z 7-58 x 5-56; FD ES cols 58-63 (Familias=61); Remuneracoes lin 68; VBP 80; Ocup 81.
"""
import os, csv, sys
import numpy as np
import openpyxl
if hasattr(sys.stdout, "reconfigure"): sys.stdout.reconfigure(encoding="utf-8")

DATA = r"C:/Users/DELL/Downloads/Material IO/Matrizes/MIP-ES-BR (2008).xlsx"
OUT  = r"C:/Users/DELL/Documents/es-insumo-produto/pesquisa/outputs"
os.makedirs(OUT, exist_ok=True)
num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0

wb = openpyxl.load_workbook(DATA, data_only=True); ws = wb["ES-BR"]
names = [str(ws.cell(r, 2).value).strip() for r in range(7, 33)]          # 26 ES
Z   = np.array([[num(ws.cell(r, c).value) for c in range(5, 57)] for r in range(7, 59)])  # 52x52
FDfamES = np.array([num(ws.cell(r, 61).value) for r in range(7, 59)])     # consumo das familias ES (52)
rem = np.array([num(ws.cell(80 if False else 68, c).value) for c in range(5, 57)])  # Remuneracoes (lin 68)
vbp = np.array([num(ws.cell(80, c).value) for c in range(5, 57)])
emp = np.array([num(ws.cell(81, c).value) for c in range(5, 57)])
wb.close()

n = 52; L = np.arange(0, 26); M = np.arange(26, 52)
x = np.where(vbp == 0, Z.sum(1) + 0, vbp); xs = np.where(x == 0, 1.0, x)
A = Z / xs[None, :]
B = np.linalg.inv(np.eye(n) - A)
Ahat = Z / xs[:, None]; G = np.linalg.inv(np.eye(n) - Ahat)   # Ghosh (para frente)

# coeficientes diretos
w_emp = emp / xs            # empregos por R$ mi
v_inc = rem / xs            # renda (remuneracoes) por R$ de producao
vL = v_inc.copy(); vL[M] = 0.0   # renda das familias DO ES (so setores do ES)

# ---- multiplicadores TIPO I (sistema bi-regional completo) ----
mp1 = B.sum(0)                              # producao
me1 = (w_emp[:, None] * B).sum(0)           # emprego total gerado / R$ mi dem. final (jobs/R$mi)
mr1 = (vL[:, None] * B).sum(0)              # renda das familias ES gerada / R$ dem. final

# ---- modelo TIPO II: fechado p/ familias DO ES ----
hh_income = rem[L].sum()
hc = FDfamES / (hh_income if hh_income else 1.0)   # coef. de consumo das familias ES
A2 = np.zeros((n + 1, n + 1)); A2[:n, :n] = A
A2[:n, n] = hc                              # coluna de consumo
A2[n, :n] = vL                              # linha de renda
B2 = np.linalg.inv(np.eye(n + 1) - A2)
mp2 = B2[:n, :n].sum(0)                     # producao tipo II (exclui linha familia)
me2 = (w_emp[:, None] * B2[:n, :n]).sum(0)
mr2 = (vL[:, None] * B2[:n, :n]).sum(0)

# geracao direta/indireta de emprego e renda por R$ 1 mi (tipo I)
emp_dir = w_emp; emp_ind = me1 - emp_dir
inc_dir = vL; inc_ind = mr1 - inc_dir

# ---- Rasmussen-Hirschman ----
bl = B.sum(0) / n / B.mean()               # para tras (Leontief)
fl = G.sum(1) / n / G.mean()               # para frente (Ghosh)

# ---- ligacoes PURAS (Guilhoto-Sonis-Hewings), setorial sobre o sistema 52 ----
# demanda final total por produto, pela identidade y = x - Z.sum(linha)
y_tot = x - Z.sum(1)
PBL = np.zeros(n); PFL = np.zeros(n)
for i in range(n):
    r = np.array([k for k in range(n) if k != i])
    A_ii = A[i, i]; A_ir = A[i, r]; A_ri = A[r, i]; A_rr = A[np.ix_(r, r)]
    Di = 1.0 / (1.0 - A_ii) if A_ii != 1 else 0.0
    Dr = np.linalg.inv(np.eye(n - 1) - A_rr)
    Xi = Di * y_tot[i]
    PBL[i] = float((Dr @ A_ri).sum() * Xi)            # impacto puro de i no resto
    PFL[i] = float(Di * (A_ir @ Dr @ y_tot[r]))       # impacto puro do resto em i
PTL = PBL + PFL
# padroniza pela media DOS SETORES DO ES (>1 => acima da media do ES => chave capixaba)
pbl_i = PBL / PBL[L].mean(); pfl_i = PFL / PFL[L].mean(); ptl_i = PTL / PTL[L].mean()

# S1 (parecer): o que a ligacao pura adiciona sobre o VBP cru?
rnk = lambda v: np.argsort(np.argsort(v))
pearson_ptl_vbp = float(np.corrcoef(PTL[L], x[L])[0, 1])
spearman_ptl_vbp = float(np.corrcoef(rnk(PTL[L]), rnk(x[L]))[0, 1])
# setores que a ligacao pura mais PROMOVE vs o ranking de VBP (sobem na conexao)
drank = rnk(PTL[L]) - rnk(x[L])   # >0 => ligacao pura o coloca acima do que o tamanho diria

# ---- coeficiente de variacao das ligacoes (grau de integracao) ----
cv_back = bl[L].std() / bl[L].mean(); cv_forw = fl[L].std() / fl[L].mean()

def br(v): return f"{v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
print("="*84)
print("C1 — ESTRUTURA PRODUTIVA DO ESPIRITO SANTO (2008) — bateria insumo-produto")
print("="*84)
print(f"setores ES=26  sistema=52  SigmaA max={A.sum(0).max():.3f}  B>=0:{bool((B>=-1e-9).all())}")
print(f"Grau de integracao (CV das ligacoes ES): para-tras {cv_back:.3f} · para-frente {cv_forw:.3f}")
print("-"*84)
print("MULTIPLICADORES (media dos 26 setores do ES):")
print(f"  producao  tipo I {mp1[L].mean():.3f}  tipo II {mp2[L].mean():.3f}")
print(f"  emprego    tipo I {me1[L].mean():.1f}  tipo II {me2[L].mean():.1f}  (empregos/R$1mi dem. final)")
print(f"  renda     tipo I {mr1[L].mean():.3f}  tipo II {mr2[L].mean():.3f}")
print("-"*84)
def top(metric, k=5, lo=False):
    o = np.argsort(metric[L]); o = o if lo else o[::-1]
    return ", ".join(f"{names[j][:22]} ({metric[L][j]:.2f})" for j in o[:k])
print("TOP-5 multiplicador de PRODUCAO (tipo I):", top(mp1))
print("TOP-5 multiplicador de EMPREGO  (tipo I):", top(me1))
print("TOP-5 multiplicador de RENDA    (tipo I):", top(mr1))
print("-"*84)
print("LIGACOES (Rasmussen-Hirschman) — SETORES-CHAVE (para-tras>1 E para-frente>1):")
chave_rh = [names[j] for j in L if bl[j] > 1 and fl[j] > 1]
print("  ", ", ".join(chave_rh) if chave_rh else "(nenhum com ambos>1)")
print("LIGACOES PURAS (GSH) — TOP-5 por PTL padronizada (centros de crescimento):")
o = np.argsort(ptl_i[L])[::-1]
for j in o[:5]:
    print(f"   {names[L[j]][:30]:30s}  PBL {pbl_i[L][j]:.2f}  PFL {pfl_i[L][j]:.2f}  PTL {ptl_i[L][j]:.2f}")
print(f"S1 — correlação PTL × VBP: Pearson {pearson_ptl_vbp:.2f} · Spearman {spearman_ptl_vbp:.2f} "
      f"(ligação pura é largamente TAMANHO; reordena pela conexão)")
prom = np.argsort(drank)[::-1][:3]
print("   setores que a ligação pura mais PROMOVE acima do que o porte diria:",
      ", ".join(f"{names[L[j]][:18]} (+{drank[j]})" for j in prom))
print("="*84)

with open(os.path.join(OUT, "es_caracterizacao_2008.csv"), "w", newline="", encoding="utf-8") as f:
    wr = csv.writer(f)
    wr.writerow(["setor","mult_prod_I","mult_prod_II","mult_emp_I","mult_emp_II",
                 "mult_renda_I","mult_renda_II","emp_dir_Rmi","emp_ind_Rmi",
                 "lig_tras","lig_frente","PBL_idx","PFL_idx","PTL_idx","chave_RH"])
    for k, j in enumerate(L):
        wr.writerow([names[k], f"{mp1[j]:.4f}", f"{mp2[j]:.4f}", f"{me1[j]*1e6:.2f}",
                     f"{me2[j]*1e6:.2f}", f"{mr1[j]:.4f}", f"{mr2[j]:.4f}",
                     f"{emp_dir[j]:.2f}", f"{emp_ind[j]:.2f}",
                     f"{bl[j]:.4f}", f"{fl[j]:.4f}", f"{pbl_i[j]:.4f}", f"{pfl_i[j]:.4f}",
                     f"{ptl_i[j]:.4f}", int(bl[j] > 1 and fl[j] > 1)])
print("salvo: es_caracterizacao_2008.csv")
