# -*- coding: utf-8 -*-
"""
01_es_br_base.py — carrega a MIP bi-regional ES x Restante do Brasil (2008) REAL
e reproduz os resultados-base do artigo (vazamento de multiplicador, spillover/
feedback, ligações de Rasmussen-Hirschman, vazamento de emprego).

Fonte: Material IO/Matrizes/MIP-ES-BR (2008).xlsx  (aba 'ES-BR', R$ milhoes)
Layout confirmado:
  linhas 7-32 = ES setores 1-26 ; 33-58 = RESTO DO BRASIL setores 1-26
  colunas 5-30 = ES (consumo intermediario) ; 31-56 = RB
  colunas 58-63 = demanda final ES (Export, Gov, ISFLSF, Familias, FBCF, VarEst)
  colunas 64-69 = demanda final RB (idem)
  linha 80 = VBP (producao total) ; linha 81 = Fator Trabalho (Ocupacoes)
"""
import os, csv
import numpy as np
import openpyxl

DATA = r"C:/Users/DELL/Downloads/Material IO/Matrizes/MIP-ES-BR (2008).xlsx"
OUT  = r"C:/Users/DELL/Documents/es-insumo-produto/pesquisa/outputs"
os.makedirs(OUT, exist_ok=True)

def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0

wb = openpyxl.load_workbook(DATA, data_only=True)
ws = wb["ES-BR"]

# rotulos de setor (col B, linhas 7-32)
names = [ws.cell(row=r, column=2).value for r in range(7, 33)]
names = [str(s).strip() for s in names]

# Z (52x52), FD (52x12), x (VBP), emprego
Z  = np.array([[num(ws.cell(row=r, column=c).value) for c in range(5, 57)] for r in range(7, 59)])
FD = np.array([[num(ws.cell(row=r, column=c).value) for c in range(58, 70)] for r in range(7, 59)])
vbp = np.array([num(ws.cell(row=80, column=c).value) for c in range(5, 57)])
emp = np.array([num(ws.cell(row=81, column=c).value) for c in range(5, 57)])
wb.close()

n = 26
L = np.arange(0, 26)    # ES
M = np.arange(26, 52)   # Resto do Brasil

# vetor de producao: usa VBP; checa contra soma de linha (Z+FD)
x_row = Z.sum(axis=1) + FD.sum(axis=1)
x = vbp.copy()
x = np.where(x == 0, x_row, x)              # fallback
consist = np.nanmax(np.abs(vbp - x_row) / np.where(x_row == 0, 1, x_row))

# coeficientes tecnicos e Leontief
xs = np.where(x == 0, 1.0, x)
A = Z / xs[np.newaxis, :]
B = np.linalg.inv(np.eye(52) - A)

# ----------------------------------------------------------------------------- #
# 1) Multiplicador de producao: retido (ES) vs vazado (RB), por setor-coluna
# ----------------------------------------------------------------------------- #
O   = B.sum(axis=0)
O_L = B[L, :].sum(axis=0)
O_M = B[M, :].sum(axis=0)
leak = O_M / np.where(O == 0, 1, O)

leak_ES = leak[L]                            # vazamento das colunas ES
mean_leak_simple = leak_ES.mean()
w_out = x[L] / x[L].sum()
mean_leak_wprod = float((leak_ES * w_out).sum())

# ----------------------------------------------------------------------------- #
# 2) Multiplicador de emprego e vazamento de emprego (colunas ES)
# ----------------------------------------------------------------------------- #
wlab = emp / xs                              # coef. de trabalho (ocupacoes / R$ mi)
E   = (wlab[:, None] * B).sum(axis=0)
E_M = (wlab[M, None] * B[M, :]).sum(axis=0)
emp_leak = E_M / np.where(E == 0, 1, E)
emp_leak_ES = emp_leak[L]

# ----------------------------------------------------------------------------- #
# 3) Ligacoes de Rasmussen-Hirschman (no sistema completo 52x52)
# ----------------------------------------------------------------------------- #
nfull = 52
Bbar = B.mean()
bl = B.sum(axis=0) / nfull / Bbar            # para tras (backward)
fl = B.sum(axis=1) / nfull / Bbar            # para frente (forward)

# ----------------------------------------------------------------------------- #
# 4) Spillover / feedback de Miller-Blair (demanda final do ES por produtos ES)
# ----------------------------------------------------------------------------- #
A_LL = A[np.ix_(L, L)]; A_LM = A[np.ix_(L, M)]
A_ML = A[np.ix_(M, L)]; A_MM = A[np.ix_(M, M)]
I_LL = np.eye(n); I_MM = np.eye(n)
inv_MM = np.linalg.inv(I_MM - A_MM)

fL    = FD[L, 0:6].sum(axis=1)              # demanda final do ES por produtos do ES
fL_RB = FD[M, 0:6].sum(axis=1)             # demanda final do ES por produtos do RB
yL_total = FD[:, 0:6].sum()                 # demanda final total do ES (todas origens)

inv_LL   = np.linalg.inv(I_LL - A_LL)                              # economia ES fechada
inv_feed = np.linalg.inv(I_LL - A_LL - A_LM @ inv_MM @ A_ML)       # com loop de feedback
xL_closed = inv_LL @ fL
xL_full   = inv_feed @ fL
xM_spill  = inv_MM @ A_ML @ xL_full         # producao induzida no RB (spillover)

inj       = float(fL.sum())                 # injecao (demanda ES por produtos ES)
spillover = float(xM_spill.sum())
feedback  = float(xL_full.sum() - xL_closed.sum())

# ----------------------------------------------------------------------------- #
# RELATORIO
# ----------------------------------------------------------------------------- #
def br(v): return f"{v:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")

print("="*70)
print("MIP-ES-BR (2008) — REPRODUCAO DOS RESULTADOS-BASE (dado real)")
print("="*70)
print(f"Consistencia VBP vs soma de linha (max erro rel.): {consist:.2%}")
print(f"Soma A (col) min/max: {A.sum(0).min():.3f} / {A.sum(0).max():.3f}  (deve ser <1)")
print(f"B>=0? {bool((B>=-1e-9).all())}   dim sistema: {B.shape}")
print("-"*70)
print("[1] VAZAMENTO DO MULTIPLICADOR DE PRODUCAO (colunas ES)")
print(f"    media simples : {mean_leak_simple:6.2%}   (artigo: 24,9%)")
print(f"    media pond.prod: {mean_leak_wprod:6.2%}")
order = np.argsort(leak_ES)
print(f"    + retido : {names[order[0]][:34]:34s} {leak_ES[order[0]]:.1%}")
print(f"    + vazado : {names[order[-1]][:34]:34s} {leak_ES[order[-1]]:.1%}")
print("-"*70)
print("[2] VAZAMENTO DE EMPREGO (colunas ES) — comparar com artigo")
alvo = {6:"Refino (art.61,6%)", 3:"Alimentos (art.56,5%)",
        5:"Madeira/papel (art.50,3%)", 10:"Metalurgia (art.49,3%)"}
for idx, lab in alvo.items():
    print(f"    {names[idx][:38]:38s} {emp_leak_ES[idx]:6.1%}   [{lab}]")
print("-"*70)
print("[4] SPILLOVER / FEEDBACK (Miller-Blair — CONVENCAO ADOTADA p/ o paper, V3)")
print(f"    Injecao f^ES (dem. final do ES por produtos do ES): R$ {br(inj)} mi")
print(f"    Spillover (producao induzida no RB)              : R$ {br(spillover)} mi")
print(f"    Feedback (volta ao ES)                           : R$ {br(feedback)} mi")
print(f"    feedback/injecao = {feedback/inj:.2%}   (tese: feedback ~ 0)   spillover/injecao = {spillover/inj:.1%}")
print(f"    [memo] demanda final total do ES (todas origens) : R$ {br(yL_total)} mi")
print("="*70)

# ----------------------------------------------------------------------------- #
# salva tabelas (versionaveis)
# ----------------------------------------------------------------------------- #
with open(os.path.join(OUT, "es_setores_resultados.csv"), "w", newline="", encoding="utf-8") as f:
    wcsv = csv.writer(f)
    wcsv.writerow(["setor", "mult_producao", "retido_ES", "vazado_RB",
                   "vazamento_prod_%", "mult_emprego", "vazamento_emprego_%",
                   "ligacao_tras", "ligacao_frente", "VBP"])
    for j in L:
        wcsv.writerow([names[j], f"{O[j]:.4f}", f"{O_L[j]:.4f}", f"{O_M[j]:.4f}",
                       f"{leak[j]*100:.2f}", f"{E[j]:.6f}", f"{emp_leak[j]*100:.2f}",
                       f"{bl[j]:.4f}", f"{fl[j]:.4f}", f"{x[j]:.1f}"])
print("salvo:", os.path.join(OUT, "es_setores_resultados.csv"))
